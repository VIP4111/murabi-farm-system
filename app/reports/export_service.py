"""
تصدير التقارير PDF/Excel (بند 22). خط عربي مُرفَق بالمشروع
(app/static/fonts/NotoNaskhArabic-Regular.ttf) عشان التصدير يشتغل بأي
بيئة نشر بدون الاعتماد على خطوط النظام.
"""
import io
import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import arabic_reshaper
from bidi.algorithm import get_display

_FONT_PATH = os.path.join(os.path.dirname(__file__), "..", "static", "fonts", "NotoNaskhArabic-Regular.ttf")
_font_registered = False


def _ensure_font():
    global _font_registered
    if not _font_registered:
        pdfmetrics.registerFont(TTFont("Arabic", _FONT_PATH))
        _font_registered = True


def ar(text) -> str:
    if text is None:
        return ""
    return get_display(arabic_reshaper.reshape(str(text)))


def build_excel(title: str, columns: list[str], rows: list[list]) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = (title or "تقرير")[:31]
    ws.sheet_view.rightToLeft = True
    # تحويل صريح لـstr() هنا (بند إضافي 165) — عناوين الأعمدة صارت
    # نصوص مترجمة (Flask-Babel LazyString)، وopenpyxl ما يقدر يكتبها
    # مباشرة بالخلية (يرفض أي نوع غير str/رقم/تاريخ صراحة).
    columns = [str(c) for c in columns]
    ws.append(columns)
    for row in rows:
        ws.append([str(cell) if cell is not None else "" for cell in row])
    for i, col in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(14, len(str(col)) + 6)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_pdf(title: str, columns: list[str], rows: list[list], subtitle: str | None = None) -> io.BytesIO:
    _ensure_font()
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    width, height = A4
    right_margin = width - 15 * mm
    left_margin = 15 * mm
    y = height - 20 * mm

    c.setFont("Arabic", 16)
    c.drawRightString(right_margin, y, ar(title))
    y -= 8 * mm
    if subtitle:
        c.setFont("Arabic", 10)
        c.drawRightString(right_margin, y, ar(subtitle))
        y -= 8 * mm

    n_cols = max(len(columns), 1)
    col_width = (right_margin - left_margin) / n_cols

    def draw_header(yy):
        c.setFont("Arabic", 9)
        for i, col in enumerate(columns):
            x = right_margin - i * col_width
            c.drawRightString(x, yy, ar(col))
        c.line(left_margin, yy - 2 * mm, right_margin, yy - 2 * mm)
        return yy - 7 * mm

    y = draw_header(y)
    c.setFont("Arabic", 8.5)
    for row in rows:
        if y < 20 * mm:
            c.showPage()
            y = height - 20 * mm
            y = draw_header(y)
            c.setFont("Arabic", 8.5)
        for i, val in enumerate(row):
            x = right_margin - i * col_width
            c.drawRightString(x, y, ar(val))
        y -= 6 * mm

    if not rows:
        c.drawRightString(right_margin, y, ar("لا يوجد بيانات لهذه الفترة."))

    c.save()
    buf.seek(0)
    return buf


def build_lot_profile_pdf(lot, rows, stats, farm_settings) -> io.BytesIO:
    """بروفايل تجاري احترافي لدفعة بيع (بند إضافي 191.3) — مستند
    عرض للمشتري المحتمل: هوية المزرعة، إحصائيات الدفعة الاستثمارية،
    وجدول تفصيلي بكل رأس (رقم، عمر، وزن، سلالة، حالة صحية عامة).
    **صفر بيانات مالية داخلية حساسة** — لا تكلفة فعلية ولا هامش ربح،
    بس الوزن والعمر والصحة (ما يهم المشتري فعلياً)."""
    _ensure_font()
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    width, height = A4
    right_margin = width - 15 * mm
    left_margin = 15 * mm
    y = height - 20 * mm

    c.setFont("Arabic", 18)
    c.drawRightString(right_margin, y, ar(lot.name))
    y -= 8 * mm
    c.setFont("Arabic", 10)
    farm_name = farm_settings.farm_name or "مراح بو علي"
    c.drawRightString(right_margin, y, ar(f"{farm_name} — {farm_settings.farm_phone or ''}"))
    y -= 10 * mm

    c.setFont("Arabic", 12)
    summary = (
        f"عدد الرؤوس: {stats['count']}   |   إجمالي الوزن: {stats['total_weight']} كجم"
        f"   |   متوسط الوزن: {stats['avg_weight']} كجم"
    )
    c.drawRightString(right_margin, y, ar(summary))
    y -= 10 * mm
    c.line(left_margin, y, right_margin, y)
    y -= 8 * mm

    columns = ["الرقم", "السلالة", "الجنس", "العمر", "الوزن (كجم)", "الحالة الصحية"]
    n_cols = len(columns)
    col_width = (right_margin - left_margin) / n_cols

    def draw_header(yy):
        c.setFont("Arabic", 9)
        for i, col in enumerate(columns):
            x = right_margin - i * col_width
            c.drawRightString(x, yy, ar(col))
        c.line(left_margin, yy - 2 * mm, right_margin, yy - 2 * mm)
        return yy - 7 * mm

    y = draw_header(y)
    c.setFont("Arabic", 8.5)
    for r in rows:
        if y < 25 * mm:
            c.showPage()
            y = height - 20 * mm
            y = draw_header(y)
            c.setFont("Arabic", 8.5)
        animal = r["animal"]
        health_state = "سليم" if r.get("open_diseases", 0) == 0 else f"{r['open_diseases']} حالة مفتوحة"
        values = [
            animal.animal_no, animal.breed or "-", animal.gender or "-",
            r["age_label"] or "-", r["weight"] or "-", health_state,
        ]
        for i, val in enumerate(values):
            x = right_margin - i * col_width
            c.drawRightString(x, y, ar(val))
        y -= 6 * mm

    y -= 6 * mm
    c.setFont("Arabic", 9)
    c.drawRightString(right_margin, y, ar("بيانات استرشادية معدَّة آلياً — تواصل معنا مباشرة للتفاصيل والمعاينة."))

    c.save()
    buf.seek(0)
    return buf


def build_invoice_pdf(finance_row, animal, farm_settings) -> io.BytesIO:
    """فاتورة بيع رسمية (بند إضافي 75) — المزرعة بائع، تصدر لمشترٍ. تخطيط
    مستند مفرد (رأس/طرفين/بند واحد/إجمالي) مو جدول تقرير، بس بنفس خط
    وأدوات build_pdf أعلاه (الخط العربي المسجَّل مرة وحدة، ودالة ar())."""
    _ensure_font()
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    width, height = A4
    right_margin = width - 20 * mm
    left_margin = 20 * mm
    y = height - 25 * mm

    c.setFont("Arabic", 18)
    c.drawRightString(right_margin, y, ar("فاتورة بيع"))
    y -= 8 * mm
    c.setFont("Arabic", 11)
    c.drawRightString(right_margin, y, ar(f"رقم الفاتورة: {finance_row.invoice_number}"))
    y -= 6 * mm
    c.drawRightString(right_margin, y, ar(f"التاريخ: {finance_row.date}"))
    y -= 12 * mm

    c.line(left_margin, y, right_margin, y)
    y -= 10 * mm

    c.setFont("Arabic", 12)
    c.drawRightString(right_margin, y, ar("البائع"))
    y -= 6 * mm
    c.setFont("Arabic", 10)
    for line in (farm_settings.farm_name, farm_settings.farm_phone, farm_settings.farm_address):
        if line:
            c.drawRightString(right_margin, y, ar(line))
            y -= 5.5 * mm
    if not (farm_settings.farm_name or farm_settings.farm_phone or farm_settings.farm_address):
        c.drawRightString(right_margin, y, ar("مراح بو علي"))
        y -= 5.5 * mm

    y -= 6 * mm
    c.setFont("Arabic", 12)
    c.drawRightString(right_margin, y, ar("المشتري"))
    y -= 6 * mm
    c.setFont("Arabic", 10)
    if finance_row.buyer_name:
        c.drawRightString(right_margin, y, ar(finance_row.buyer_name))
        y -= 5.5 * mm
    if finance_row.buyer_phone:
        c.drawRightString(right_margin, y, ar(finance_row.buyer_phone))
        y -= 5.5 * mm
    if not (finance_row.buyer_name or finance_row.buyer_phone):
        c.drawRightString(right_margin, y, ar("غير مسجَّل"))
        y -= 5.5 * mm

    y -= 12 * mm
    col_item = right_margin
    col_amount = left_margin + 30 * mm
    c.setFont("Arabic", 10)
    c.drawRightString(col_item, y, ar("البيان"))
    c.drawRightString(col_amount, y, ar("المبلغ"))
    y -= 3 * mm
    c.line(left_margin, y, right_margin, y)
    y -= 8 * mm

    item_label = f"بيع رأس رقم {animal.animal_no}" if animal else (finance_row.item or "بيع")
    c.setFont("Arabic", 10)
    c.drawRightString(col_item, y, ar(item_label))
    c.drawRightString(col_amount, y, ar(f"{finance_row.amount:,.2f}"))
    y -= 6 * mm
    c.line(left_margin, y, right_margin, y)
    y -= 10 * mm

    c.setFont("Arabic", 13)
    c.drawRightString(right_margin, y, ar(f"الإجمالي: {finance_row.amount:,.2f}"))

    if finance_row.description:
        y -= 12 * mm
        c.setFont("Arabic", 9)
        c.drawRightString(right_margin, y, ar(f"ملاحظات: {finance_row.description}"))

    # رمز QR بأسلوب "فاتورة" المرحلة الأولى (بند إضافي 184) — بس لو
    # صاحب الحلال سجّل رقماً ضريبياً فعلياً بالإعدادات؛ فاضي = بدون رمز
    # إطلاقاً (نفس فلسفة كل ميزة اختيارية بالمشروع: صفر إعداد = صفر أثر).
    from app.core.zatca_service import invoice_qr_image
    from datetime import datetime as _dt
    qr_buf = invoice_qr_image(
        farm_settings=farm_settings, invoice_total=finance_row.amount,
        timestamp=_dt.combine(finance_row.date, _dt.min.time()),
    )
    if qr_buf:
        from reportlab.lib.utils import ImageReader
        qr_size = 28 * mm
        c.drawImage(ImageReader(qr_buf), left_margin, 15 * mm, width=qr_size, height=qr_size)
        c.setFont("Arabic", 7)
        c.drawString(left_margin, 12 * mm, ar("QR مبسّط (المرحلة الأولى) — ليس فاتورة ضريبية معتمدة رسمياً"))

    c.save()
    buf.seek(0)
    return buf


_ARABIC_MONTHS = {
    1: "يناير", 2: "فبراير", 3: "مارس", 4: "أبريل", 5: "مايو", 6: "يونيو",
    7: "يوليو", 8: "أغسطس", 9: "سبتمبر", 10: "أكتوبر", 11: "نوفمبر", 12: "ديسمبر",
}


def build_payroll_receipt_pdf(payroll, farm_settings) -> io.BytesIO:
    """مسير راتب شهر واحد (بند إضافي 242) — نظام الرواتب العام (بخلاف
    وصل "موظف الشهر" الأبسط، بند 240): يفصّل الراتب الأساسي + المكافأة
    - كل سطر خصم بسببه - = الصافي المستحق، بطلبك الصريح."""
    _ensure_font()
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    width, height = A4
    right_margin = width - 20 * mm
    left_margin = 20 * mm
    y = height - 25 * mm

    c.setFont("Arabic", 18)
    c.drawRightString(right_margin, y, ar("مسير راتب الشهر"))
    y -= 10 * mm
    c.setFont("Arabic", 10)
    c.drawRightString(right_margin, y, ar(f"التاريخ: {payroll.confirmed_at.date() if payroll.confirmed_at else ''}"))
    y -= 12 * mm

    c.line(left_margin, y, right_margin, y)
    y -= 10 * mm

    c.setFont("Arabic", 12)
    c.drawRightString(right_margin, y, ar("صاحب العمل"))
    y -= 6 * mm
    c.setFont("Arabic", 10)
    c.drawRightString(right_margin, y, ar(farm_settings.farm_name or "مراح بو علي"))
    y -= 5.5 * mm
    if farm_settings.owner_national_id:
        c.drawRightString(right_margin, y, ar(f"رقم الهوية: {farm_settings.owner_national_id}"))
        y -= 5.5 * mm
    if farm_settings.farm_phone:
        c.drawRightString(right_margin, y, ar(f"رقم الجوال: {farm_settings.farm_phone}"))
        y -= 5.5 * mm

    y -= 8 * mm
    c.setFont("Arabic", 12)
    c.drawRightString(right_margin, y, ar("بيانات العامل"))
    y -= 6 * mm
    c.setFont("Arabic", 10)
    c.drawRightString(right_margin, y, ar(f"اسم العامل: {payroll.user.name}"))
    y -= 5.5 * mm
    if payroll.user.nationality:
        c.drawRightString(right_margin, y, ar(f"الجنسية: {payroll.user.nationality}"))
        y -= 5.5 * mm
    if payroll.user.passport_number:
        c.drawRightString(right_margin, y, ar(f"رقم الجواز: {payroll.user.passport_number}"))
        y -= 5.5 * mm
    if payroll.user.border_number:
        c.drawRightString(right_margin, y, ar(f"رقم الحدود: {payroll.user.border_number}"))
        y -= 5.5 * mm
    c.drawRightString(right_margin, y, ar(f"فترة الراتب: {_ARABIC_MONTHS.get(payroll.month, payroll.month)} {payroll.year}"))
    y -= 5.5 * mm
    if payroll.user.payment_method == "تحويل بنكي":
        # حوالة (بند إضافي 244) — بطلبك: "من المحوّل ومن مستلم الحوالة"
        # صريحين. المحوّل = صاحب الحلال (مذكور فوق أصلاً)، والمستلم
        # حقل مستقل قابل للاستبدال كل شهر (Payroll.recipient_name).
        recipient = payroll.recipient_name or payroll.user.name
        c.drawRightString(right_margin, y, ar(f"طريقة الدفع: تحويل بنكي — من {farm_settings.farm_name or 'صاحب الحلال'} إلى {recipient}"))
        y -= 5.5 * mm
    elif payroll.user.payment_method:
        c.drawRightString(right_margin, y, ar(f"طريقة الدفع: {payroll.user.payment_method}"))
        y -= 5.5 * mm

    y -= 10 * mm
    col_item = right_margin
    col_amount = left_margin + 30 * mm
    c.setFont("Arabic", 10)
    c.drawRightString(col_item, y, ar("البيان"))
    c.drawRightString(col_amount, y, ar("المبلغ"))
    y -= 3 * mm
    c.line(left_margin, y, right_margin, y)
    y -= 7 * mm

    c.setFont("Arabic", 10)
    c.drawRightString(col_item, y, ar("الراتب الأساسي"))
    c.drawRightString(col_amount, y, ar(f"{payroll.base_salary:,.2f}"))
    y -= 6 * mm
    if payroll.bonus_amount:
        c.drawRightString(col_item, y, ar("المكافأة"))
        c.drawRightString(col_amount, y, ar(f"{payroll.bonus_amount:,.2f}"))
        y -= 6 * mm
    for d in payroll.deductions:
        label = f"خصم — {d.reason}" if d.reason else "خصم"
        c.drawRightString(col_item, y, ar(label))
        c.drawRightString(col_amount, y, ar(f"-{d.amount:,.2f}"))
        y -= 6 * mm

    y -= 2 * mm
    c.line(left_margin, y, right_margin, y)
    y -= 10 * mm

    c.setFont("Arabic", 13)
    c.drawRightString(right_margin, y, ar(f"الصافي المستحق: {payroll.net_amount:,.2f}"))
    y -= 10 * mm
    c.setFont("Arabic", 9)
    c.drawRightString(right_margin, y, ar(f"معتمَد من: {payroll.confirmed_by.name if payroll.confirmed_by else '-'}"))

    y -= 30 * mm
    c.setFont("Arabic", 10)
    c.drawRightString(right_margin, y, ar("توقيع العامل: ......................"))
    y -= 12 * mm
    c.drawRightString(right_margin, y, ar("توقيع صاحب الحلال: ......................"))

    c.save()
    buf.seek(0)
    return buf
